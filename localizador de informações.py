import os
import openpyxl
import docx2txt
import win32com.client

# Especifique o caminho do arquivo XLSX
caminho_arquivo = "C:\\Users\\Lucas.Beltrao\\Seuarquivo.xlxs"

# Carregue o arquivo XLSX
workbook = openpyxl.load_workbook(caminho_arquivo)

# Escolha uma planilha (por exemplo, a primeira planilha)
sheet = workbook.active

# Crie uma lista para armazenar os dados
dados = []

# Percorra as linhas da planilha
for row in sheet.iter_rows(values_only=True):
    # Verifique se pelo menos uma célula na linha não está vazia
    if any(row):
        dados.append(row)

# Agora, 'dados' contém as informações do arquivo XLSX como uma lista de listas
# Cada lista interna representa uma linha na planilha

# Exemplo: Imprima os dados
for coluna in dados:
    print(coluna)

# Transforme a lista de listas em uma lista de strings
dados_em_strings = [str(item) for row in dados for item in row]

print(dados_em_strings)

# Agora, 'dados_em_strings' conterá todos os elementos da lista 'dados' como strings
# Cada elemento da lista corresponderá a um item da lista 'dados'

# Defina a pasta raiz onde deseja iniciar a pesquisa
pasta_raiz = "X:\\pastaraiz"

# Defina a lista de strings que você deseja localizar nos arquivos
strings_a_encontrar = (dados_em_strings)

# Crie uma instância do aplicativo Microsoft Word
word = win32com.client.Dispatch("Word.Application")

# Lista para armazenar as correspondências encontradas
correspondencias = []

# Percorra todas as pastas e subpastas
for pasta_atual, subpastas, arquivos in os.walk(pasta_raiz):
    for arquivo in arquivos:
        caminho_completo = os.path.join(pasta_atual, arquivo)

        # Verifique se o arquivo é .docx
        if arquivo.endswith('.docx'):
            # Use docx2txt para extrair o texto do arquivo do Word
            try:
                texto = docx2txt.process(caminho_completo)
                
                # Verifique se alguma das strings da lista está presente no texto
                for string in strings_a_encontrar:
                    if string in texto:
                        correspondencias.append((caminho_completo, string, texto))
                        print(f"CM ou Desvio {string} encontrado no arquivo {caminho_completo}")
            except Exception as e:
                print(0)

        # Verifique se o arquivo é .doc
        elif arquivo.endswith('.doc'):
            try:
                # Abra o arquivo .doc no Microsoft Word
                doc = word.Documents.Open(caminho_completo)
                
                # Obtenha o conteúdo do arquivo
                texto = doc.Content.Text
                
                # Verifique se alguma das strings da lista está presente no texto
                for string in strings_a_encontrar:
                    if string in texto:
                        correspondencias.append((caminho_completo, string, texto))
                        print(f"CM ou Desvio {string} encontrado no arquivo {caminho_completo}")
                        
                # Feche o arquivo .doc
                doc.Close()
            except Exception as e:
                print(0)

# Feche o aplicativo Microsoft Word
word.Quit()

import openpyxl

# Crie um novo arquivo Excel
planilha = openpyxl.Workbook()
folha = planilha.active

# Adicionar cabeçalhos à planilha
folha.append(["Arquivo", "String Encontrada"])

# Adicionar as correspondências à planilha
for caminho, string, texto in correspondencias:
    folha.append([caminho, string])  # Remova 'texto' desta linha

# Especifique o diretório onde deseja salvar o arquivo e o nome do arquivo
diretorio = "C:\\diretorio"
nome_arquivo = "correspondencias.xlsx"

# Combine o diretório e o nome do arquivo para criar o caminho completo
caminho_completo = os.path.join(diretorio, nome_arquivo)

# Salvar a planilha em um arquivo Excel no diretório especificado
planilha.save(caminho_completo)

# Fechar a planilha
planilha.close()

print(f"O arquivo Excel '{caminho_completo}' foi criado com as correspondências.")
